import sampler
import requests
import random
import datetime
import numpy as np
import collections
from statistics import mean
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from sampler.configuation import Connector


class Statistics(Connector):

    def __init__(self):
        super().__init__()
        

    def random_numbers_pandas(self, popdata, popsize, samplesize):
        mastersampledata = [[]]
        i = 0
        while i < self.details["number_sampled"]: #len(mastersampledata) < sampler.numberofsampled: 
            if len(mastersampledata[-1]) <= (samplesize - (samplesize * .02)) or len(mastersampledata[-1]) >= (samplesize + (samplesize * .02)):
                print("Sample size does not meet criteria. Deleting...")
                del mastersampledata[-1]
                i=i-1
            else:
                print("Random Sample Accepted")
            if i == -1:
                i=0
            mastersampledata.append(popdata)
            mastersampledata[i]['randNumCol'] = np.random.randint(1, 6, mastersampledata[i].shape[0])
            randomnumber = int(round(random.random()*(popsize/samplesize),0))
            mastersampledata[i] = mastersampledata[i][mastersampledata[i]['randNumCol'] == randomnumber]
            #mastersampledata[i].rename(columns={mastersampledata[i].columns[1]: "id", mastersampledata[i].columns[2]: "data"}, inplace = True)
            i+=1 
        samplenumber = len(mastersampledata)
        lenlists = []
        for values in mastersampledata:
            lenlists.append(len(values))
        avg = mean(lenlists)
        print(mastersampledata)
        print(str(samplenumber) + " samples obtained. Average outlets per sample: " + str(avg))
        return mastersampledata


    def drop_duplicates(self, popdata, sampledata):
        return pd.concat([sampledata, popdata]).drop_duplicates(keep=False)


    def average_aggregation(self, samplegroup, testgroup):
        averages = []
        for group in testgroup:
            pass


    

    def attritionmodeling(masterdata, outlets):
        attrpos = masterdata[0].index('result')
        outletpos = masterdata[0].index('outlet_id')
        attr = []
        result = []
        for z in range(len(outlets)):
            attr.append([])
            for i in range(len(outlets[z])):
                for j in range(len(masterdata[1])):
                    if outlets[z][i] == masterdata[1][j][outletpos]:
                        attr[z].append(int(masterdata[1][j][attrpos]))
            result.append(collections.Counter(attr[z]))
        return result

    def attritionmodelingdict(results, outlets):
        attr = []
        result = []
        for i in range(len(outlets)):
            attr.append([])
            for j in range(len(outlets[i])):
                attr[i].append(str(results[outlets[i][j]]))
            result.append(collections.Counter(attr[i]))
        return result
        
    def dictkeylist(dictionarylist):
        keylist = []
        for dictionary in dictionarylist:
            for key in dictionary:
                keylist.append(key)
        return keylist

    def importExcel(cleankeys, testdata, sampledata):
        wb = Workbook()
        ws = wb.active
        ws.title = "SourceData"
        for y in range(len(cleankeys)):
            ws.cell(row=1, column=y+1).value = cleankeys[y]
        for x in range(len(testdata[0])):
            try:
                ws.cell(row=2, column=x+1).value = testdata[0][str(ws.cell(row=1, column=x+1).value)]
            except:
                continue
        for x in range(len(sampledata)):
            for y in range(len(sampledata[x])):
                try:
                    ws.cell(row=x+3, column=y+1).value = sampledata[x][str(ws.cell(row=1, column=y+1).value)]
                except:
                    continue
        wb.save(filename = sampler.filepath)
        
                

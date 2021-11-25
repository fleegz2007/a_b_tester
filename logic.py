import requests
import config
import random
import datetime
import collections
from statistics import mean
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

token = config.foundrytoken

def query_foundry_sql(query, token, branch='master', base_url='https://one-palantir.1dc.com'):

#Queries the dataproxy query API with spark SQL.
#Example: query_foundry_sql("SELECT * FROM `/path/to/dataset` Limit 5000", "ey...")
#Args:
#    query: the sql query
#    branch: the branch of the dataset / query
#
#Returns: (columns, data) tuple. data contains the data matrix, columns the list of columns
#Can be converted to a pandas Dataframe:
#pd.DataFrame(data, columns)
    response = requests.post(f"{base_url}foundry-data-proxy/api/dataproxy/queryWithFallbacks",
                         headers={'Authorization': f'Bearer {token}'},
                         params={'fallbackBranchIds': [branch]},
                         json={'query': query})
    response.raise_for_status()
    json = response.json()
    columns = [e['name'] for e in json['foundrySchema']['fieldSchemaList']]
    return columns, json['rows']

def devqueryexcel():
    groupings = []
    wb = load_workbook(config.excelfile)
    ws = wb.active
    maxcol = ws.max_col
    mincol = ws.min_col
    maxrow = ws.max_row
    for x in range(2,maxrow):
        groupings.append([])
        for y in range(mincol,maxcol):
            groupings[x].append(ws.cell(row=x, column=y))
    return groupings
    
def obtain_testgroup(testtable):
    query = config.rapid_deposit_attritiontest
    samplegroup = query_foundry_sql(query, token, branch='master', base_url='https://one-palantir.1dc.com/')
    return samplegroup

def obtain_population(populationtable):
    query = config.rapid_deposit_attritionsample
    population = query_foundry_sql(query, token, branch='master', base_url='https://one-palantir.1dc.com/')
    return population

def random_numbers(popdata, popsize, samplesize):
    mastersampledata = [[]]
    i = 0
    while i < config.numberofsampled: #len(mastersampledata) < config.numberofsampled: 
        if len(mastersampledata[-1]) <= (samplesize - (samplesize * .02)) or len(mastersampledata[-1]) >= (samplesize + (samplesize * .02)):
            #print("Sample size does not meet criteria. Deleting...")
            del mastersampledata[-1]
            i=i-1
        if i == -1:
            i=0
        mastersampledata.append([])
        randlist = []
        for j in range(popsize):
            randlist.append(int(round(random.random()*(popsize/samplesize),0)))
        randomnumber = int(round(random.random()*(popsize/samplesize),0))
        #print("Random Number Selected: " + str(randomnumber))
        for j in range(popsize):
            if randlist[j] == randomnumber:
                mastersampledata[i].append(popdata[j])
        #print("Sample obtained. Sample size: " + str(len(mastersampledata[i])))
        i+=1 
    samplenumber = len(mastersampledata)
    lenlists = []
    for values in mastersampledata:
        lenlists.append(len(values))
    avg = mean(lenlists)
    print(str(samplenumber) + " samples obtained. Average outlets per sample: " + str(avg))
    return mastersampledata

def attritionmodeling(masterdata, outlets):
    attrpos = masterdata[0].index('result')
    outletpos = masterdata[0].index('outlet_id')
    attr = []
    result = []
    for z in range(len(outlets)):
        attr.append([])
        for i in range(len(outlets[z])):
            for j in range(len(masterdata[1])):
                if outlets[z][i] == masterdata[1][j][outletpos]:
                    attr[z].append(int(masterdata[1][j][attrpos]))
        result.append(collections.Counter(attr[z]))
    return result

def attritionmodelingdict(results, outlets):
    attr = []
    result = []
    for i in range(len(outlets)):
        attr.append([])
        for j in range(len(outlets[i])):
            attr[i].append(str(results[outlets[i][j]]))
        result.append(collections.Counter(attr[i]))
    return result
    
def dictkeylist(dictionarylist):
    keylist = []
    for dictionary in dictionarylist:
        for key in dictionary:
            keylist.append(key)
    return keylist

def importExcel(cleankeys, testdata, sampledata):
    wb = Workbook()
    ws = wb.active
    ws.title = "SourceData"
    for y in range(len(cleankeys)):
        ws.cell(row=1, column=y+1).value = cleankeys[y]
    for x in range(len(testdata[0])):
        try:
            ws.cell(row=2, column=x+1).value = testdata[0][str(ws.cell(row=1, column=x+1).value)]
        except:
            continue
    for x in range(len(sampledata)):
        for y in range(len(sampledata[x])):
            try:
                ws.cell(row=x+3, column=y+1).value = sampledata[x][str(ws.cell(row=1, column=y+1).value)]
            except:
                continue
    wb.save(filename = config.filepath)
    
            
